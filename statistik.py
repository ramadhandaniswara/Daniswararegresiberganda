# Import Library yang diperlukan
import pandas as pd
import argparse
from math import pow, sqrt
from decimal import Decimal
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# Buat parameter argumen input script
parser = argparse.ArgumentParser(description='Script penghitung Regresi Linear Ganda.')
# Argumen input name file
parser.add_argument('-i','--input', help='Input file name',required=True)
# Argumen output nama file
parser.add_argument('-o','--output',help='Output file name', required=True)
# Parsing Argumen
args = parser.parse_args()
# Load data dari excel
workbook = load_workbook(args.input)
# Pilih worksheet yang active
worksheet = workbook.active
# Definisikan maksimal kolom pada data
maks_kolom = worksheet.max_column
# Definisikan maksimal baris pada data
maks_baris = worksheet.max_row
# Buat Penampungan Array
x1, x2, y, y_topi = [], [], [], []
# Buat Variabel Penampungan
sigma_x1 = 0
sigma_x2 = 0
sigma_y = 0
sigma_x1y = 0
sigma_x2y = 0
sigma_x1x2 = 0
sigma_x1pow2 = 0
sigma_x2pow2 = 0
sigma_ypow2 = 0
# Jumlah baris data
n = maks_baris - 1
# Tentukan koefisien
k = 3
# buat perulangan untuk mengambil data yang ada
for i in range(2, maks_baris + 1):
    cell_object = worksheet.cell(row = i, column = 2) 
    x1.append(cell_object.value)

for i in range(2, maks_baris + 1):
    cell_object = worksheet.cell(row = i, column = 4)
    x2.append(cell_object.value)

for i in range(2, maks_baris + 1):
    cell_object = worksheet.cell(row = i, column = 5)
    y.append(cell_object.value)
# Hitung data dengan rumus yang ditentukan
for x in range(0, n):
    # Hitung Total x1
    sigma_x1 += Decimal(x1[x])
    # Hitung Total x2
    sigma_x2 += x2[x]
    # Hitung Total y
    sigma_y += y[x]
    # Hitung x1.y
    x1y = Decimal(x1[x]) * y[x]
    sigma_x1y += x1y
    # Hitung x2.y
    x2y = x2[x] * y[x]
    sigma_x2y += x2y
    # Hitung x1.x2
    x1x2 = Decimal(x1[x]) * x2[x]
    sigma_x1x2 += x1x2
    # Hitung x1^2
    x1pow2 = pow(Decimal(x1[x]), 2)
    sigma_x1pow2 += x1pow2
    # Hitung x1^2
    x2pow2 = pow(x2[x], 2)
    sigma_x2pow2 += x2pow2
    # Hitung y^2
    ypow2 = pow(y[x], 2)
    sigma_ypow2 += ypow2

# Cari b1, b2, a
# b1
b1_1 = sigma_x2pow2 * float(sigma_x1y) - float(sigma_x1x2) * sigma_x2y
b1_2 = sigma_x1pow2 * sigma_x2pow2 - pow(sigma_x1x2, 2)
b1 = b1_1 / b1_2
# b2
b2_1 = sigma_x1pow2 * sigma_x2y - float(sigma_x1x2) * float(sigma_x1y)
b2_2 = sigma_x1pow2 * sigma_x2pow2 - pow(sigma_x1x2, 2)
b2 = b2_1 / b2_2
# a
a = sigma_y / n - b1 * (float(sigma_x1) / n) - b2 * (sigma_x2 / 32)
# Buat perulangan untuk menghitung Y topi
for k in range(0, n):
    # Hitung Y topi
    ytopi = a + (b1 * float(x1[k]) + (b2 * float(x2[k])))
    # Masukkan data ke variabel array y_topi
    y_topi.append(ytopi)
# Hitung Standar Error
sy12_1 = sigma_ypow2 - (b1 * float(sigma_x1y)) - (b2 * sigma_x2y)
sy12_2 = n - k
sy12 = sqrt(sy12_1) / sqrt(sy12_2)
# Hitung korelasi linier berganda
rx1y = float(sigma_x1y) / sqrt(sigma_x1pow2 * sigma_ypow2)
rx2y = sigma_x2y / sqrt(sigma_x2pow2 * sigma_ypow2)
# Hitung Besar Pengaruh Terhadap y
x1_persen = pow(rx1y, 2) * 100
x2_persen = pow(rx2y, 2) * 100
# Outputkan data di file dan sheet yang baru
writer = pd.ExcelWriter(args.output, engine='openpyxl')
wb = writer.book
# definisikan data yang akan ditulis bagian 1
df = pd.DataFrame({
    'X1': x1,
    'X2': x2,
    'Y': y,
    "Y'": y_topi
})
# definisikan data yang akan ditulis bagian 2
df1 = pd.DataFrame({
    'estimasi parameter (a)': [a],
    'estimasi parameter (b1)': [b1],
    'estimasi parameter (b2)': [b2],
    'standar error (sy12)': [sy12],
    'korelasi(rx1y)': [rx1y],
    'korelasi(rx2y)': [rx2y],
    'besar pengaruh y(x1)': [x1_persen],
    'besar pengaruh y(x2)': [x2_persen]
})
# gabung data bagian 1 dan bagian 2
data_output = pd.concat([df, df1], axis=1)
# tulis data ke excel
data_output.to_excel(writer, index=False)
# simpan file yang baru
wb.save(args.output)
# Outputkan ringkasan data
print('- persamaan regresi linear ganda')
print("y = {}(a) + {}(x1) + {}(x2)".format(a, b1, b2))
print('')
print('- estimasi standar error')
print('sy12 = {}'.format(sy12))
print('')
print('- korelasi regresi linier berganda')
print('rx1y = {}'.format(rx1y))
print('rx2y = {}'.format(rx2y))
print('')
print('- besar pengaruh terhadap y')
print('x1 = {}%'.format(x1_persen))
print('x2 = {}%'.format(x2_persen))
print('')
print('- made with love by Kelompok 2 -')